import sys
import json
import logging
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional, Set

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PySide6.QtCore import Qt, QPointF, Signal, QMimeData, QRectF
from PySide6.QtGui import QAction, QColor, QBrush, QPen, QDrag, QPainter, QWheelEvent
from PySide6.QtWidgets import (
    QApplication,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFormLayout,
    QGraphicsDropShadowEffect,
    QGridLayout,
    QGroupBox,
    QHeaderView,
    QHBoxLayout,
    QInputDialog,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QSizePolicy,
    QAbstractSpinBox,
    QSpinBox,
    QDoubleSpinBox,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QVBoxLayout,
    QWidget,
    QGraphicsItem,
    QGraphicsRectItem,
    QGraphicsScene,
    QGraphicsSimpleTextItem,
    QGraphicsView,
)

APP_DIR = Path(__file__).resolve().parent
DATA_DIR = APP_DIR / "data"
LOG_DIR = APP_DIR / "logs"
STYLE_PATH = APP_DIR / "styles" / "app_style.qss"
DB_PATH = DATA_DIR / "parts.db"
TEMPLATE_PATH = DATA_DIR / "parts_upload_template.xlsx"
RAW_EXPORT_DEFAULT = DATA_DIR / "parts_raw_data.xlsx"
CANVAS_SAVE_PATH = DATA_DIR / "breaker_canvas_layout.json"
CANVAS_EXPORT_PATH = DATA_DIR / "breaker_canvas_layout.xlsx"
BREAKER_TEMPLATE_PATH = DATA_DIR / "breaker_templates.json"

DATA_DIR.mkdir(exist_ok=True)
LOG_DIR.mkdir(exist_ok=True)


# ------------------------------------------------------------
# Logging
# ------------------------------------------------------------
def setup_logging() -> logging.Logger:
    logger = logging.getLogger("electrical_capacity_app")
    if logger.handlers:
        return logger

    logger.setLevel(logging.DEBUG)
    formatter = logging.Formatter(
        "%(asctime)s | %(levelname)s | %(name)s | %(funcName)s | %(message)s"
    )

    file_handler = logging.FileHandler(LOG_DIR / "app.log", encoding="utf-8")
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(formatter)

    stream_handler = logging.StreamHandler(sys.stdout)
    stream_handler.setLevel(logging.INFO)
    stream_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(stream_handler)
    logger.debug("Logging initialized")
    return logger


logger = setup_logging()


def apply_card_shadow(widget: QWidget):
    effect = QGraphicsDropShadowEffect(widget)
    effect.setBlurRadius(14)
    effect.setOffset(0, 2)
    effect.setColor(QColor(15, 23, 42, 32))
    widget.setGraphicsEffect(effect)


DEFAULT_BREAKER_TEMPLATES = [
    {"label": "ELCB", "prefix": "ELCB"},
    {"label": "MCCB", "prefix": "MCCB"},
    {"label": "CP", "prefix": "CP"},
]


def load_breaker_templates() -> List[dict]:
    templates = list(DEFAULT_BREAKER_TEMPLATES)
    if BREAKER_TEMPLATE_PATH.exists():
        try:
            saved = json.loads(BREAKER_TEMPLATE_PATH.read_text(encoding="utf-8"))
            if isinstance(saved, list):
                for item in saved:
                    prefix = str(item.get("prefix", "")).strip().upper()
                    label = str(item.get("label", prefix)).strip()
                    if prefix and not any(t["prefix"] == prefix for t in templates):
                        templates.append({"label": label or prefix, "prefix": prefix})
        except Exception:
            logger.exception("Failed to load breaker templates")
    return templates


def save_breaker_templates(templates: List[dict]):
    payload = []
    for item in templates:
        prefix = str(item.get("prefix", "")).strip().upper()
        label = str(item.get("label", prefix)).strip()
        if prefix:
            payload.append({"label": label or prefix, "prefix": prefix})
    BREAKER_TEMPLATE_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


# ------------------------------------------------------------
# Data layer
# ------------------------------------------------------------
@dataclass
class PartRecord:
    part_no: str
    part_name: str
    category: str
    voltage_v: float
    current_a: float
    power_w: float
    phase: str
    power_factor: float
    recommended_breaker_a: float
    note: str = ""


class DatabaseManager:
    def __init__(self, db_path: Path):
        self.db_path = db_path
        logger.debug("DatabaseManager init: %s", db_path)
        self._initialize()

    def _connect(self):
        return sqlite3.connect(self.db_path)

    def _initialize(self):
        logger.info("Initializing database")
        with self._connect() as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS parts (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    part_no TEXT NOT NULL UNIQUE,
                    part_name TEXT NOT NULL,
                    category TEXT,
                    voltage_v REAL NOT NULL,
                    current_a REAL NOT NULL,
                    power_w REAL NOT NULL,
                    phase TEXT NOT NULL,
                    power_factor REAL NOT NULL,
                    recommended_breaker_a REAL NOT NULL,
                    note TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
            conn.commit()
        logger.info("Database ready")

    def upsert_part(self, part: PartRecord):
        logger.info("Upserting part: %s / %s", part.part_no, part.part_name)
        with self._connect() as conn:
            conn.execute(
                """
                INSERT INTO parts (
                    part_no, part_name, category, voltage_v, current_a, power_w,
                    phase, power_factor, recommended_breaker_a, note
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(part_no) DO UPDATE SET
                    part_name=excluded.part_name,
                    category=excluded.category,
                    voltage_v=excluded.voltage_v,
                    current_a=excluded.current_a,
                    power_w=excluded.power_w,
                    phase=excluded.phase,
                    power_factor=excluded.power_factor,
                    recommended_breaker_a=excluded.recommended_breaker_a,
                    note=excluded.note
                """,
                (
                    part.part_no,
                    part.part_name,
                    part.category,
                    part.voltage_v,
                    part.current_a,
                    part.power_w,
                    part.phase,
                    part.power_factor,
                    part.recommended_breaker_a,
                    part.note,
                ),
            )
            conn.commit()

    def get_all_parts(self) -> pd.DataFrame:
        with self._connect() as conn:
            return pd.read_sql_query(
                "SELECT part_no, part_name, category, voltage_v, current_a, power_w, phase, power_factor, recommended_breaker_a, note, created_at FROM parts ORDER BY part_no",
                conn,
            )

    def get_part_by_no(self, part_no: str) -> Optional[pd.Series]:
        with self._connect() as conn:
            df = pd.read_sql_query(
                "SELECT * FROM parts WHERE part_no = ?",
                conn,
                params=(part_no,),
            )
        if df.empty:
            return None
        return df.iloc[0]

    def get_part_nos(self) -> List[str]:
        with self._connect() as conn:
            rows = conn.execute("SELECT part_no FROM parts ORDER BY part_no").fetchall()
        return [r[0] for r in rows]

    def bulk_upsert_from_dataframe(self, df: pd.DataFrame) -> int:
        required_columns = [
            "part_no",
            "part_name",
            "category",
            "voltage_v",
            "current_a",
            "power_w",
            "phase",
            "power_factor",
            "recommended_breaker_a",
            "note",
        ]
        missing = [c for c in required_columns if c not in df.columns]
        if missing:
            raise ValueError(f"엑셀 양식 컬럼이 누락되었습니다: {missing}")

        count = 0
        for _, row in df.iterrows():
            part = PartRecord(
                part_no=str(row["part_no"]).strip(),
                part_name=str(row["part_name"]).strip(),
                category=str(row.get("category", "")).strip(),
                voltage_v=float(row["voltage_v"]),
                current_a=float(row["current_a"]),
                power_w=float(row["power_w"]),
                phase=str(row["phase"]).strip(),
                power_factor=float(row["power_factor"]),
                recommended_breaker_a=float(row["recommended_breaker_a"]),
                note=str(row.get("note", "")).strip(),
            )
            self.upsert_part(part)
            count += 1
        return count


class ExcelManager:
    TEMPLATE_COLUMNS = {
        "part_no": ["MTR-001"],
        "part_name": ["Main Motor"],
        "category": ["Motor"],
        "voltage_v": [220],
        "current_a": [5.2],
        "power_w": [1144],
        "phase": ["1P"],
        "power_factor": [0.95],
        "recommended_breaker_a": [15],
        "note": ["Sample row. Delete after checking format."],
    }

    @staticmethod
    def create_template(path: Path):
        df = pd.DataFrame(ExcelManager.TEMPLATE_COLUMNS)
        df.to_excel(path, index=False)

    @staticmethod
    def load_excel(path: Path) -> pd.DataFrame:
        return pd.read_excel(path)

    @staticmethod
    def export_dataframe(df: pd.DataFrame, path: Path):
        df.to_excel(path, index=False)


class LoadCalculator:
    BREAKER_STANDARDS = [5, 10, 15, 20, 30, 40, 50, 60, 75, 100, 125, 150, 175, 200, 225, 250, 300, 400]

    @staticmethod
    def calculate_total(part_row: pd.Series, quantity: int, safety_factor: float = 1.25) -> dict:
        total_current = float(part_row["current_a"]) * quantity
        total_power = float(part_row["power_w"]) * quantity
        safety_current = total_current * safety_factor
        suggested_breaker = LoadCalculator.select_breaker(safety_current)
        return {
            "part_no": part_row["part_no"],
            "part_name": part_row["part_name"],
            "quantity": quantity,
            "voltage_v": float(part_row["voltage_v"]),
            "phase": part_row["phase"],
            "unit_current_a": float(part_row["current_a"]),
            "unit_power_w": float(part_row["power_w"]),
            "total_current_a": round(total_current, 2),
            "total_power_w": round(total_power, 2),
            "safety_factor": round(safety_factor, 2),
            "safety_current_a": round(safety_current, 2),
            "suggested_breaker_a": suggested_breaker,
        }

    @staticmethod
    def select_breaker(required_current: float) -> int:
        if required_current < 0:
            required_current = 0
        for size in LoadCalculator.BREAKER_STANDARDS:
            if size >= required_current:
                return size
        return LoadCalculator.BREAKER_STANDARDS[-1]


# ------------------------------------------------------------
# Parts / Calculator tabs
# ------------------------------------------------------------
class PartsTab(QWidget):
    def __init__(self, db: DatabaseManager):
        super().__init__()
        self.db = db
        self._build_ui()
        self.refresh_table()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        form_group = QGroupBox("파트 등록")
        form_layout = QGridLayout()

        self.part_no_edit = QLineEdit()
        self.part_name_edit = QLineEdit()
        self.category_edit = QLineEdit()
        self.voltage_spin = QDoubleSpinBox()
        self.voltage_spin.setRange(0, 100000)
        self.voltage_spin.setDecimals(2)
        self.voltage_spin.setValue(220)
        self.voltage_spin.setButtonSymbols(QAbstractSpinBox.NoButtons)

        self.current_spin = QDoubleSpinBox()
        self.current_spin.setRange(0, 100000)
        self.current_spin.setDecimals(2)
        self.current_spin.setButtonSymbols(QAbstractSpinBox.NoButtons)

        self.power_spin = QDoubleSpinBox()
        self.power_spin.setRange(0, 10000000)
        self.power_spin.setDecimals(2)
        self.power_spin.setButtonSymbols(QAbstractSpinBox.NoButtons)

        self.phase_combo = QComboBox()
        self.phase_combo.addItems(["1P", "3P", "DC"])

        self.pf_spin = QDoubleSpinBox()
        self.pf_spin.setRange(0, 1)
        self.pf_spin.setSingleStep(0.01)
        self.pf_spin.setValue(0.95)
        self.pf_spin.setButtonSymbols(QAbstractSpinBox.NoButtons)

        self.breaker_spin = QDoubleSpinBox()
        self.breaker_spin.setRange(0, 10000)
        self.breaker_spin.setDecimals(0)
        self.breaker_spin.setButtonSymbols(QAbstractSpinBox.NoButtons)

        self.note_edit = QLineEdit()

        widgets = [
            ("Part No", self.part_no_edit),
            ("Part Name", self.part_name_edit),
            ("Category", self.category_edit),
            ("Voltage (V)", self.voltage_spin),
            ("Current (A)", self.current_spin),
            ("Power (W)", self.power_spin),
            ("Phase", self.phase_combo),
            ("Power Factor", self.pf_spin),
            ("Breaker (A)", self.breaker_spin),
            ("Note", self.note_edit),
        ]

        for i, (label, widget) in enumerate(widgets):
            row, col = divmod(i, 2)
            form_layout.addWidget(QLabel(label), row * 2, col)
            form_layout.addWidget(widget, row * 2 + 1, col)

        form_group.setLayout(form_layout)
        apply_card_shadow(form_group)

        button_row = QHBoxLayout()
        self.save_btn = QPushButton("파트 저장")
        self.template_btn = QPushButton("엑셀 양식 다운로드")
        self.import_btn = QPushButton("엑셀 일괄 등록")
        self.export_btn = QPushButton("등록 파트 Raw Data 다운로드")
        self.refresh_btn = QPushButton("목록 새로고침")
        for btn in [self.save_btn, self.template_btn, self.import_btn, self.export_btn, self.refresh_btn]:
            button_row.addWidget(btn)
        button_row.addStretch()

        self.table = QTableWidget()
        self.table.setColumnCount(10)
        self.table.setHorizontalHeaderLabels(
            [
                "Part No", "Name", "Category", "Voltage(V)", "Current(A)",
                "Power(W)", "Phase", "PF", "Breaker(A)", "Note",
            ]
        )
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        apply_card_shadow(self.table)

        layout.addWidget(form_group)
        layout.addLayout(button_row)
        layout.addWidget(self.table)

        self.save_btn.clicked.connect(self.save_part)
        self.template_btn.clicked.connect(self.download_template)
        self.import_btn.clicked.connect(self.import_excel)
        self.export_btn.clicked.connect(self.export_raw_data)
        self.refresh_btn.clicked.connect(self.refresh_table)

    def save_part(self):
        try:
            part = PartRecord(
                part_no=self.part_no_edit.text().strip(),
                part_name=self.part_name_edit.text().strip(),
                category=self.category_edit.text().strip(),
                voltage_v=self.voltage_spin.value(),
                current_a=self.current_spin.value(),
                power_w=self.power_spin.value(),
                phase=self.phase_combo.currentText(),
                power_factor=self.pf_spin.value(),
                recommended_breaker_a=self.breaker_spin.value(),
                note=self.note_edit.text().strip(),
            )
            if not part.part_no or not part.part_name:
                raise ValueError("Part No와 Part Name은 필수입니다.")
            self.db.upsert_part(part)
            self.refresh_table()
            QMessageBox.information(self, "완료", "파트가 저장되었습니다.")
        except Exception as e:
            logger.exception("Failed to save part")
            QMessageBox.critical(self, "오류", str(e))

    def refresh_table(self):
        df = self.db.get_all_parts()
        self.table.setRowCount(len(df))
        columns = [
            "part_no", "part_name", "category", "voltage_v", "current_a",
            "power_w", "phase", "power_factor", "recommended_breaker_a", "note",
        ]
        for row_idx, (_, row) in enumerate(df.iterrows()):
            for col_idx, col in enumerate(columns):
                self.table.setItem(row_idx, col_idx, QTableWidgetItem(str(row[col])))

    def download_template(self):
        try:
            save_path, _ = QFileDialog.getSaveFileName(
                self, "엑셀 양식 저장", str(TEMPLATE_PATH), "Excel Files (*.xlsx)"
            )
            if not save_path:
                return
            ExcelManager.create_template(Path(save_path))
            QMessageBox.information(self, "완료", f"양식이 저장되었습니다.\n{save_path}")
        except Exception as e:
            logger.exception("Failed to create template")
            QMessageBox.critical(self, "오류", str(e))

    def import_excel(self):
        try:
            file_path, _ = QFileDialog.getOpenFileName(
                self, "엑셀 파일 선택", str(APP_DIR), "Excel Files (*.xlsx *.xls)"
            )
            if not file_path:
                return
            df = ExcelManager.load_excel(Path(file_path))
            count = self.db.bulk_upsert_from_dataframe(df)
            self.refresh_table()
            QMessageBox.information(self, "완료", f"{count}개 파트가 등록/업데이트되었습니다.")
        except Exception as e:
            logger.exception("Failed to import excel")
            QMessageBox.critical(self, "오류", str(e))

    def export_raw_data(self):
        try:
            save_path, _ = QFileDialog.getSaveFileName(
                self, "등록 파트 Raw Data 저장", str(RAW_EXPORT_DEFAULT), "Excel Files (*.xlsx)"
            )
            if not save_path:
                return
            df = self.db.get_all_parts()
            ExcelManager.export_dataframe(df, Path(save_path))
            QMessageBox.information(self, "완료", f"Raw Data가 저장되었습니다.\n{save_path}")
        except Exception as e:
            logger.exception("Failed to export raw data")
            QMessageBox.critical(self, "오류", str(e))


class CalcTab(QWidget):
    def __init__(self, db: DatabaseManager):
        super().__init__()
        self.db = db
        self._build_ui()
        self.reload_part_numbers()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        input_group = QGroupBox("부하 계산 / 차단기 선정")
        form = QFormLayout()

        self.part_combo = QComboBox()
        self.qty_spin = QSpinBox()
        self.qty_spin.setRange(1, 100000)
        self.qty_spin.setValue(1)
        self.qty_spin.setButtonSymbols(QAbstractSpinBox.NoButtons)
        self.safety_factor_spin = QDoubleSpinBox()
        self.safety_factor_spin.setRange(1.0, 5.0)
        self.safety_factor_spin.setSingleStep(0.05)
        self.safety_factor_spin.setValue(1.25)
        self.safety_factor_spin.setButtonSymbols(QAbstractSpinBox.NoButtons)
        self.calc_btn = QPushButton("계산 실행")
        self.reload_btn = QPushButton("파트 목록 다시 불러오기")

        form.addRow("등록 파트", self.part_combo)
        form.addRow("수량", self.qty_spin)
        form.addRow("안전율", self.safety_factor_spin)
        form.addRow(self.calc_btn, self.reload_btn)
        input_group.setLayout(form)
        apply_card_shadow(input_group)

        result_group = QGroupBox("계산 결과")
        result_layout = QGridLayout()
        self.result_labels = {}
        result_fields = [
            "part_no", "part_name", "quantity", "voltage_v", "phase",
            "unit_current_a", "unit_power_w", "total_current_a", "total_power_w",
            "safety_factor", "safety_current_a", "suggested_breaker_a",
        ]
        for i, key in enumerate(result_fields):
            title = QLabel(key)
            value = QLabel("-")
            self.result_labels[key] = value
            r, c = divmod(i, 2)
            result_layout.addWidget(title, r, c * 2)
            result_layout.addWidget(value, r, c * 2 + 1)
        result_group.setLayout(result_layout)
        apply_card_shadow(result_group)

        layout.addWidget(input_group)
        layout.addWidget(result_group)
        layout.addStretch()

        self.calc_btn.clicked.connect(self.run_calculation)
        self.reload_btn.clicked.connect(self.reload_part_numbers)

    def reload_part_numbers(self):
        self.part_combo.clear()
        self.part_combo.addItems(self.db.get_part_nos())

    def run_calculation(self):
        try:
            part_no = self.part_combo.currentText().strip()
            if not part_no:
                raise ValueError("먼저 등록된 파트를 선택해 주세요.")
            part_row = self.db.get_part_by_no(part_no)
            if part_row is None:
                raise ValueError("선택한 파트를 찾을 수 없습니다.")
            result = LoadCalculator.calculate_total(
                part_row, self.qty_spin.value(), self.safety_factor_spin.value()
            )
            for key, value in result.items():
                self.result_labels[key].setText(str(value))
        except Exception as e:
            logger.exception("Failed to calculate")
            QMessageBox.critical(self, "오류", str(e))


# ------------------------------------------------------------
# Canvas drag sources
# ------------------------------------------------------------
class DraggablePartListWidget(QListWidget):
    def __init__(self, db: DatabaseManager):
        super().__init__()
        self.db = db
        self._all_parts: list[dict] = []
        self._search_text = ""
        self.setDragEnabled(True)
        self.setAlternatingRowColors(True)
        self.refresh_parts()

    def refresh_parts(self):
        df = self.db.get_all_parts()
        self._all_parts = []
        for _, row in df.iterrows():
            self._all_parts.append({
                "part_no": str(row["part_no"]),
                "part_name": str(row["part_name"]),
                "current_a": row["current_a"],
                "power_w": row["power_w"],
            })
        self.apply_filter(self._search_text)

    def apply_filter(self, search_text: str = ""):
        self._search_text = (search_text or "").strip().lower()
        self.clear()

        for row in self._all_parts:
            haystack = f"{row['part_no']} {row['part_name']}".lower()
            if self._search_text and self._search_text not in haystack:
                continue

            text = f"{row['part_no']} | {row['part_name']} | {row['current_a']}A | {row['power_w']}W"
            item = QListWidgetItem(text)
            item.setData(Qt.UserRole, row['part_no'])
            self.addItem(item)

    def startDrag(self, supportedActions):
        item = self.currentItem()
        if item is None:
            return
        part_no = item.data(Qt.UserRole)
        if not part_no:
            return
        drag = QDrag(self)
        mime = QMimeData()
        mime.setText(str(part_no))
        mime.setData("application/x-item-type", b"part")
        drag.setMimeData(mime)
        drag.exec(Qt.CopyAction)


class BreakerTemplateListWidget(QListWidget):
    def __init__(self):
        super().__init__()
        self.setDragEnabled(True)
        self.setAlternatingRowColors(True)
        self.refresh_templates(load_breaker_templates())

    def refresh_templates(self, templates: List[dict]):
        self.clear()
        for template in templates:
            prefix = str(template.get("prefix", "")).strip().upper()
            label = str(template.get("label", prefix)).strip()
            if not prefix:
                continue
            item = QListWidgetItem(label)
            item.setData(Qt.UserRole, prefix)
            item.setToolTip(f"드래그하여 {prefix} 차단기 생성")
            self.addItem(item)

    def startDrag(self, supportedActions):
        item = self.currentItem()
        if item is None:
            return
        prefix = item.data(Qt.UserRole)
        if not prefix:
            return
        drag = QDrag(self)
        mime = QMimeData()
        mime.setText(str(prefix))
        mime.setData("application/x-item-type", b"breaker")
        drag.setMimeData(mime)
        drag.exec(Qt.CopyAction)


class BreakerTemplateDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("차단기 템플릿 추가")
        layout = QVBoxLayout(self)
        form = QFormLayout()

        self.type_combo = QComboBox()
        self.type_combo.addItems(["ELCB", "MCCB", "CP", "CUSTOM"])
        self.label_edit = QLineEdit()
        self.prefix_edit = QLineEdit()
        self.type_combo.currentTextChanged.connect(self._sync_defaults)
        self._sync_defaults(self.type_combo.currentText())

        form.addRow("기본 타입", self.type_combo)
        form.addRow("표시 이름", self.label_edit)
        form.addRow("접두어", self.prefix_edit)
        layout.addLayout(form)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _sync_defaults(self, text: str):
        if text != "CUSTOM":
            self.label_edit.setText(text)
            self.prefix_edit.setText(text)

    def values(self) -> tuple[str, str]:
        label = self.label_edit.text().strip()
        prefix = self.prefix_edit.text().strip().upper()
        return label, prefix


class BreakerSettingsDialog(QDialog):
    def __init__(self, breaker_name: str, safety_factor: float, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"차단기 설정 - {breaker_name}")
        layout = QVBoxLayout(self)
        form = QFormLayout()

        self.name_edit = QLineEdit(breaker_name)
        self.safety_spin = QDoubleSpinBox()
        self.safety_spin.setRange(1.0, 5.0)
        self.safety_spin.setSingleStep(0.05)
        self.safety_spin.setValue(safety_factor)

        form.addRow("차단기 이름", self.name_edit)
        form.addRow("안전율", self.safety_spin)
        layout.addLayout(form)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def values(self) -> tuple[str, float]:
        return self.name_edit.text().strip(), self.safety_spin.value()


class LoadPartItem(QGraphicsRectItem):
    DEFAULT_WIDTH = 200
    DEFAULT_HEIGHT = 68
    MIN_WIDTH = 160
    MIN_HEIGHT = 60
    HANDLE_SIZE = 12

    def __init__(self, scene_ref: 'BreakerCanvasScene', parent_breaker: 'BreakerItem', part_row: pd.Series,
                 quantity: int = 1, width: float | None = None, height: float | None = None):
        self._width = float(width or self.DEFAULT_WIDTH)
        self._height = float(height or self.DEFAULT_HEIGHT)
        super().__init__(0, 0, self._width, self._height)
        self.scene_ref = scene_ref
        self.parent_breaker = parent_breaker
        self.part_no = str(part_row['part_no'])
        self.part_name = str(part_row['part_name'])
        self.unit_current = float(part_row['current_a'])
        self.unit_power = float(part_row['power_w'])
        self.quantity = quantity
        self._resizing = False
        self._resize_origin = QPointF()
        self._start_size = (self._width, self._height)

        self.setBrush(QBrush(QColor("#fff8ef")))
        self.setPen(QPen(QColor("#d9c2a6"), 1.2))
        self.setFlag(QGraphicsItem.ItemIsMovable, True)
        self.setFlag(QGraphicsItem.ItemIsSelectable, True)
        self.setFlag(QGraphicsItem.ItemSendsGeometryChanges, True)
        self.setAcceptHoverEvents(True)
        self.setZValue(3)

        self.label = QGraphicsSimpleTextItem(self)
        self.label.setBrush(QBrush(QColor("#4a3f35")))
        self.update_display()

    def _resize_handle_rect(self) -> QRectF:
        return QRectF(self._width - self.HANDLE_SIZE - 4, self._height - self.HANDLE_SIZE - 4,
                      self.HANDLE_SIZE, self.HANDLE_SIZE)

    def _apply_rect(self):
        self.setRect(0, 0, self._width, self._height)

    def update_display(self):
        total_current = self.unit_current * self.quantity
        total_power = self.unit_power * self.quantity
        self.label.setText(
            f"{self.part_no}\n{self.part_name}\nQty {self.quantity} | {total_current:.2f}A | {total_power:.0f}W"
        )
        self.label.setPos(10, 8)

    def paint(self, painter: QPainter, option, widget=None):
        super().paint(painter, option, widget)
        painter.setBrush(QBrush(QColor("#cbd5e1")))
        painter.setPen(QPen(QColor("#94a3b8"), 1.0))
        painter.drawRect(self._resize_handle_rect())

    def itemChange(self, change, value):
        if change == QGraphicsItem.ItemPositionHasChanged:
            self.scene_ref.notify_layout_changed()
            if self.parent_breaker:
                self.parent_breaker.refresh_recursive()
        return super().itemChange(change, value)

    def mousePressEvent(self, event):
        if self._resize_handle_rect().contains(event.pos()):
            self._resizing = True
            self._resize_origin = event.scenePos()
            self._start_size = (self._width, self._height)
            event.accept()
            return
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self._resizing:
            delta = event.scenePos() - self._resize_origin
            self.prepareGeometryChange()
            self._width = max(self.MIN_WIDTH, self._start_size[0] + delta.x())
            self._height = max(self.MIN_HEIGHT, self._start_size[1] + delta.y())
            self._apply_rect()
            self.update_display()
            self.scene_ref.notify_layout_changed()
            event.accept()
            return
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        if self._resizing:
            self._resizing = False
            self.scene_ref.notify_layout_changed()
            event.accept()
            return
        super().mouseReleaseEvent(event)

    def mouseDoubleClickEvent(self, event):
        menu = QMessageBox()
        menu.setWindowTitle("부하 파트 편집")
        menu.setText(f"{self.part_no} / {self.part_name}")
        qty_btn = menu.addButton("수량 변경", QMessageBox.AcceptRole)
        del_btn = menu.addButton("삭제", QMessageBox.DestructiveRole)
        menu.addButton("취소", QMessageBox.RejectRole)
        menu.exec()

        clicked = menu.clickedButton()
        if clicked == qty_btn:
            qty, ok = QInputDialog.getInt(
                None,
                "수량 변경",
                f"{self.part_no} 수량",
                self.quantity,
                1,
                100000,
                1,
            )
            if ok:
                self.quantity = qty
                self.update_display()
                if self.parent_breaker:
                    self.parent_breaker.refresh_recursive()
                self.scene_ref.notify_layout_changed()
        elif clicked == del_btn:
            self.scene_ref.delete_item(self)
        super().mouseDoubleClickEvent(event)

    def to_dict(self) -> dict:
        pos = self.scenePos()
        return {
            "part_no": self.part_no,
            "quantity": self.quantity,
            "x": round(pos.x(), 2),
            "y": round(pos.y(), 2),
            "width": round(self._width, 2),
            "height": round(self._height, 2),
        }


class BreakerItem(QGraphicsRectItem):
    DEFAULT_WIDTH = 280
    DEFAULT_HEIGHT = 120
    MIN_WIDTH = 220
    MIN_HEIGHT = 120
    HANDLE_SIZE = 14

    def __init__(self, scene_ref: 'BreakerCanvasScene', name: str, pos_x: float, pos_y: float,
                 safety_factor: float = 1.25, parent_breaker: Optional['BreakerItem'] = None,
                 is_top_level: bool = False, width: float | None = None, height: float | None = None,
                 breaker_type: str = "MCCB"):
        self._width = float(width or self.DEFAULT_WIDTH)
        self._height = float(height or self.DEFAULT_HEIGHT)
        super().__init__(0, 0, self._width, self._height)
        self.scene_ref = scene_ref
        self.name = name
        self.breaker_type = (breaker_type or "MCCB").upper()
        self.safety_factor = safety_factor
        self.parent_breaker = parent_breaker
        self.child_breakers: List['BreakerItem'] = []
        self.load_items: List[LoadPartItem] = []
        self.is_top_level = is_top_level
        self._resizing = False
        self._resize_origin = QPointF()
        self._start_size = (self._width, self._height)

        self.setPos(pos_x, pos_y)
        self.setBrush(QBrush(QColor("#eef4ff")))
        self.setPen(QPen(QColor("#9cb4de"), 1.6))
        self.setFlag(QGraphicsItem.ItemIsMovable, True)
        self.setFlag(QGraphicsItem.ItemIsSelectable, True)
        self.setFlag(QGraphicsItem.ItemSendsGeometryChanges, True)
        self.setAcceptDrops(False)
        self.setZValue(2)

        self.title_text = QGraphicsSimpleTextItem(self)
        self.title_text.setBrush(QBrush(QColor("#334155")))
        self.summary_text = QGraphicsSimpleTextItem(self)
        self.summary_text.setBrush(QBrush(QColor("#475569")))
        self.hint_text = QGraphicsSimpleTextItem("더블클릭: 이름/안전율 설정", self)
        self.hint_text.setBrush(QBrush(QColor("#64748b")))
        self._layout_text_items()
        self.update_summary()

    def _resize_handle_rect(self) -> QRectF:
        return QRectF(self._width - self.HANDLE_SIZE - 6, self._height - self.HANDLE_SIZE - 6,
                      self.HANDLE_SIZE, self.HANDLE_SIZE)

    def _layout_text_items(self):
        self.title_text.setPos(12, 10)
        self.summary_text.setPos(12, 38)
        self.hint_text.setPos(12, max(90, self._height - 24))

    def _apply_rect(self):
        self.setRect(0, 0, self._width, self._height)
        self._layout_text_items()

    def paint(self, painter: QPainter, option, widget=None):
        super().paint(painter, option, widget)
        painter.setBrush(QBrush(QColor("#cbd5e1")))
        painter.setPen(QPen(QColor("#94a3b8"), 1.0))
        painter.drawRect(self._resize_handle_rect())

    def itemChange(self, change, value):
        if change == QGraphicsItem.ItemPositionHasChanged:
            self.scene_ref.notify_layout_changed()
        return super().itemChange(change, value)

    def mousePressEvent(self, event):
        if self._resize_handle_rect().contains(event.pos()):
            self._resizing = True
            self._resize_origin = event.scenePos()
            self._start_size = (self._width, self._height)
            event.accept()
            return
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self._resizing:
            delta = event.scenePos() - self._resize_origin
            self.prepareGeometryChange()
            self._width = max(self.MIN_WIDTH, self._start_size[0] + delta.x())
            self._height = max(self.MIN_HEIGHT, self._start_size[1] + delta.y())
            self._apply_rect()
            self.update_summary()
            self.scene_ref.notify_layout_changed()
            event.accept()
            return
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        if self._resizing:
            self._resizing = False
            self.scene_ref.notify_layout_changed()
            event.accept()
            return
        super().mouseReleaseEvent(event)

    def mouseDoubleClickEvent(self, event):
        dialog = BreakerSettingsDialog(self.name, self.safety_factor)
        if dialog.exec() == QDialog.Accepted:
            new_name, new_safety_factor = dialog.values()
            if new_name:
                self.name = new_name
                inferred = new_name.split("-")[0].strip().upper()
                if inferred:
                    self.breaker_type = inferred
            self.safety_factor = new_safety_factor
            self.refresh_recursive()
            self.scene_ref.notify_layout_changed()
        super().mouseDoubleClickEvent(event)

    def is_ancestor_of(self, other: 'BreakerItem') -> bool:
        cur = other.parent_breaker
        while cur is not None:
            if cur is self:
                return True
            cur = cur.parent_breaker
        return False

    def can_accept_breaker(self, child: 'BreakerItem') -> bool:
        if child is self:
            return False
        if child.is_ancestor_of(self):
            return False
        return True

    def set_parent_breaker(self, new_parent: Optional['BreakerItem']):
        old_parent = self.parent_breaker
        if old_parent is new_parent:
            return
        if old_parent is not None and self in old_parent.child_breakers:
            old_parent.child_breakers.remove(self)
            old_parent.refresh_recursive()
        self.parent_breaker = new_parent
        if new_parent is not None and self not in new_parent.child_breakers:
            new_parent.child_breakers.append(self)
        self.refresh_recursive()

    def add_part(self, part_row: pd.Series, quantity: int = 1, drop_pos: Optional[QPointF] = None,
                 width: float | None = None, height: float | None = None):
        load_item = LoadPartItem(self.scene_ref, self, part_row, quantity, width=width, height=height)
        self.scene_ref.addItem(load_item)
        self.load_items.append(load_item)
        if drop_pos is None:
            drop_pos = QPointF(self.scenePos().x() + 40, self.scenePos().y() + self._height + 60)
        load_item.setPos(drop_pos)
        self.refresh_recursive()
        self.scene_ref.notify_layout_changed()
        return load_item

    def add_child_breaker(self, child_breaker: 'BreakerItem') -> bool:
        if not self.can_accept_breaker(child_breaker):
            QMessageBox.warning(None, "연결 불가", "자기 자신 또는 상위 차단기를 하위에 연결할 수 없습니다.")
            return False
        child_breaker.set_parent_breaker(self)
        self.refresh_recursive()
        self.scene_ref.notify_layout_changed()
        return True

    def remove_load_item(self, load_item: LoadPartItem):
        if load_item in self.load_items:
            self.load_items.remove(load_item)
            self.refresh_recursive()

    def get_total_current(self, visited: Optional[Set[int]] = None) -> float:
        if visited is None:
            visited = set()
        obj_id = id(self)
        if obj_id in visited:
            logger.warning("Cycle detected in get_total_current: %s", self.name)
            return 0.0
        visited.add(obj_id)
        total = sum(item.unit_current * item.quantity for item in self.load_items)
        for child in self.child_breakers:
            total += child.get_total_current(visited)
        return total

    def get_total_power(self, visited: Optional[Set[int]] = None) -> float:
        if visited is None:
            visited = set()
        obj_id = id(self)
        if obj_id in visited:
            logger.warning("Cycle detected in get_total_power: %s", self.name)
            return 0.0
        visited.add(obj_id)
        total = sum(item.unit_power * item.quantity for item in self.load_items)
        for child in self.child_breakers:
            total += child.get_total_power(visited)
        return total

    def get_total_load_count(self, visited: Optional[Set[int]] = None) -> int:
        if visited is None:
            visited = set()
        obj_id = id(self)
        if obj_id in visited:
            logger.warning("Cycle detected in get_total_load_count: %s", self.name)
            return 0
        visited.add(obj_id)
        total = len(self.load_items)
        for child in self.child_breakers:
            total += child.get_total_load_count(visited)
        return total

    def suggested_breaker(self) -> int:
        return LoadCalculator.select_breaker(self.get_total_current() * self.safety_factor)

    def update_summary(self):
        self.title_text.setText(self.name)
        self.summary_text.setText(
            f"안전율 {self.safety_factor:.2f} | 부하 {self.get_total_load_count()}개\n"
            f"합계 {self.get_total_current():.2f}A / {self.get_total_power():.0f}W | 추천 {self.suggested_breaker()}A"
        )
        self._layout_text_items()

    def refresh_recursive(self):
        self.update_summary()
        if self.parent_breaker is not None:
            self.parent_breaker.refresh_recursive()

    def contains_scene_pos(self, scene_pos: QPointF) -> bool:
        return self.sceneBoundingRect().contains(scene_pos)

    def to_dict(self) -> dict:
        pos = self.scenePos()
        return {
            "name": self.name,
            "x": round(pos.x(), 2),
            "y": round(pos.y(), 2),
            "width": round(self._width, 2),
            "height": round(self._height, 2),
            "safety_factor": self.safety_factor,
            "breaker_type": self.breaker_type,
            "is_top_level": self.is_top_level,
            "loads": [item.to_dict() for item in self.load_items],
            "children": [child.to_dict() for child in self.child_breakers],
        }


# ------------------------------------------------------------
# Canvas scene / view
# ------------------------------------------------------------
# Canvas scene / view
# ------------------------------------------------------------
class BreakerCanvasScene(QGraphicsScene):
    layoutChanged = Signal()

    def __init__(self, db: DatabaseManager):
        super().__init__()
        self.db = db
        self.top_breaker: Optional[BreakerItem] = None
        self.all_breakers: List[BreakerItem] = []
        self.connection_lines = []
        self.breaker_seq = 1
        self.setSceneRect(-50000, -50000, 100000, 100000)
        self.setBackgroundBrush(QBrush(QColor("#f8fafc")))

    def ensure_visible_scene_area(self):
        if not self.all_breakers:
            return
        bounds = self.itemsBoundingRect().adjusted(-1200, -1200, 1200, 1200)
        scene_rect = self.sceneRect()
        if not scene_rect.contains(bounds):
            self.setSceneRect(scene_rect.united(bounds))

    def new_breaker_name(self, breaker_type: str = "MCCB") -> str:
        prefix = (breaker_type or "MCCB").upper()
        while True:
            name = f"{prefix}-{self.breaker_seq:02d}"
            self.breaker_seq += 1
            if not any(b.name == name for b in self.all_breakers):
                return name

    def get_top_breakers(self) -> List[BreakerItem]:
        return [b for b in self.all_breakers if b.parent_breaker is None]

    def add_top_breaker(self, name: Optional[str] = None, x: float = 120, y: float = 80, breaker_type: str = "MCCB") -> Optional[BreakerItem]:
        if self.top_breaker is not None:
            QMessageBox.information(None, "안내", "최상위 차단기는 1개만 생성할 수 있습니다.")
            return None
        resolved_name = name or f"MAIN-{(breaker_type or 'MCCB').upper()}"
        breaker = BreakerItem(self, resolved_name, x, y, is_top_level=True, breaker_type=breaker_type)
        self.addItem(breaker)
        self.all_breakers.append(breaker)
        self.top_breaker = breaker
        self.notify_layout_changed()
        return breaker

    def create_child_breaker(self, parent_breaker: BreakerItem, pos: Optional[QPointF] = None,
                             name: Optional[str] = None, breaker_type: str = "MCCB") -> BreakerItem:
        if pos is None:
            pos = QPointF(parent_breaker.scenePos().x() + 380, parent_breaker.scenePos().y() + 220)
        breaker = BreakerItem(
            self,
            name or self.new_breaker_name(breaker_type),
            pos.x(),
            pos.y(),
            parent_breaker=parent_breaker,
            breaker_type=breaker_type,
        )
        self.addItem(breaker)
        self.all_breakers.append(breaker)
        parent_breaker.child_breakers.append(breaker)
        parent_breaker.refresh_recursive()
        self.notify_layout_changed()
        return breaker

    def find_breaker_at(self, scene_pos: QPointF) -> Optional[BreakerItem]:
        items = self.items(scene_pos)
        for item in items:
            if isinstance(item, BreakerItem):
                return item
            if isinstance(item, QGraphicsSimpleTextItem) and isinstance(item.parentItem(), BreakerItem):
                return item.parentItem()
        return None

    def update_connection_lines(self):
        for line in self.connection_lines:
            self.removeItem(line)
        self.connection_lines.clear()

        pen = QPen(QColor("#b8c4d6"), 1.2)
        for breaker in self.all_breakers:
            rect = breaker.rect()
            start_x = breaker.scenePos().x() + rect.width() / 2
            start_y = breaker.scenePos().y() + rect.height()

            for load_item in breaker.load_items:
                load_rect = load_item.rect()
                end_x = load_item.scenePos().x() + load_rect.width() / 2
                end_y = load_item.scenePos().y()
                line = self.addLine(start_x, start_y, end_x, end_y, pen)
                line.setZValue(-10)
                self.connection_lines.append(line)

            for child in breaker.child_breakers:
                child_rect = child.rect()
                end_x = child.scenePos().x() + child_rect.width() / 2
                end_y = child.scenePos().y()
                line = self.addLine(start_x, start_y, end_x, end_y, pen)
                line.setZValue(-10)
                self.connection_lines.append(line)

    def notify_layout_changed(self):
        for breaker in self.all_breakers:
            breaker.update_summary()
        self.update_connection_lines()
        self.ensure_visible_scene_area()
        self.layoutChanged.emit()

    def delete_item(self, item):
        if isinstance(item, LoadPartItem):
            if item.parent_breaker:
                item.parent_breaker.remove_load_item(item)
            self.removeItem(item)
            self.notify_layout_changed()
            return

        if isinstance(item, BreakerItem):
            if item.is_top_level:
                QMessageBox.information(None, "안내", "최상위 차단기는 삭제할 수 없습니다.")
                return
            for child in list(item.child_breakers):
                child.set_parent_breaker(item.parent_breaker)
            for load in list(item.load_items):
                item.remove_load_item(load)
                self.removeItem(load)
            if item.parent_breaker and item in item.parent_breaker.child_breakers:
                item.parent_breaker.child_breakers.remove(item)
                item.parent_breaker.refresh_recursive()
            if item in self.all_breakers:
                self.all_breakers.remove(item)
            self.removeItem(item)
            self.notify_layout_changed()

    def delete_selected_items(self):
        selected = list(self.selectedItems())
        if not selected:
            QMessageBox.information(None, "안내", "삭제할 항목을 먼저 선택하세요.")
            return
        for item in selected:
            if isinstance(item, (LoadPartItem, BreakerItem)):
                self.delete_item(item)

    def handle_drop(self, mime, scene_pos: QPointF):
        item_type = bytes(mime.data("application/x-item-type")).decode(errors="ignore")
        target_breaker = self.find_breaker_at(scene_pos)

        if item_type == "part":
            part_no = mime.text()
            if not target_breaker:
                QMessageBox.warning(None, "배치 실패", "차단기 카드 위에 드롭해야 합니다.")
                return
            part_row = self.db.get_part_by_no(part_no)
            if part_row is None:
                QMessageBox.warning(None, "배치 실패", "해당 파트를 찾을 수 없습니다.")
                return
            target_breaker.add_part(part_row, 1, scene_pos)
            return

        if item_type == "breaker":
            if not target_breaker:
                QMessageBox.warning(None, "배치 실패", "하위 차단기는 기존 차단기 위에 드롭해야 합니다.")
                return
            breaker_type = (mime.text() or "MCCB").strip().upper()
            self.create_child_breaker(
                target_breaker,
                QPointF(scene_pos.x(), scene_pos.y() + 160),
                breaker_type=breaker_type,
            )

    def to_dict(self) -> dict:
        if self.top_breaker is None:
            return {"top_breaker": None}
        return {"top_breaker": self.top_breaker.to_dict()}

    def clear_canvas(self):
        self.clear()
        self.connection_lines.clear()
        self.all_breakers.clear()
        self.top_breaker = None

    def save_to_file(self, path: Path):
        path.write_text(json.dumps(self.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")

    def export_to_excel(self, path: Path):
        breaker_rows = []
        load_rows = []
        tree_rows = []

        def walk_summary(breaker: BreakerItem, parent_name: str = ""):
            breaker_rows.append({
                "차단기 이름": breaker.name,
                "상위 차단기": parent_name,
                "차단기 종류": breaker.breaker_type,
                "안전율": round(breaker.safety_factor, 2),
                "전체 부하 수": breaker.get_total_load_count(),
                "합계 전류(A)": round(breaker.get_total_current(), 2),
                "합계 전력(W)": round(breaker.get_total_power(), 2),
                "차단기 용량(A)": breaker.suggested_breaker(),
            })

            for load in breaker.load_items:
                load_rows.append({
                    "소속 차단기": breaker.name,
                    "부하 Part No": load.part_no,
                    "부하 이름": load.part_name,
                    "수량": load.quantity,
                    "단위 전류(A)": round(load.unit_current, 4),
                    "합산 전류(A)": round(load.unit_current * load.quantity, 4),
                    "단위 전력(W)": round(load.unit_power, 2),
                    "합산 전력(W)": round(load.unit_power * load.quantity, 2),
                })

            for child in breaker.child_breakers:
                walk_summary(child, breaker.name)

        def walk_tree(breaker: BreakerItem, level: int = 0, parent_name: str = ""):
            tree_rows.append({
                "레벨": level,
                "구분": "차단기",
                "이름": breaker.name,
                "상위": parent_name,
                "차단기 종류": breaker.breaker_type,
                "수량": 1,
                "단위 전류(A)": "",
                "합산 전류(A)": round(breaker.get_total_current(), 2),
                "단위 전력(W)": "",
                "합산 전력(W)": round(breaker.get_total_power(), 2),
                "안전율": round(breaker.safety_factor, 2),
                "차단기 용량(A)": breaker.suggested_breaker(),
                "비고": "최상위 차단기" if level == 0 else "",
            })

            for child_breaker in breaker.child_breakers:
                walk_tree(child_breaker, level + 1, breaker.name)

            for load in breaker.load_items:
                tree_rows.append({
                    "레벨": level + 1,
                    "구분": "부하",
                    "이름": f"{load.part_no} / {load.part_name}",
                    "상위": breaker.name,
                    "차단기 종류": "",
                    "수량": load.quantity,
                    "단위 전류(A)": round(load.unit_current, 4),
                    "합산 전류(A)": round(load.unit_current * load.quantity, 4),
                    "단위 전력(W)": round(load.unit_power, 2),
                    "합산 전력(W)": round(load.unit_power * load.quantity, 2),
                    "안전율": "",
                    "차단기 용량(A)": "",
                    "비고": "",
                })

        if self.top_breaker:
            walk_summary(self.top_breaker)
            walk_tree(self.top_breaker)

        df_tree = pd.DataFrame(tree_rows)
        df_breakers = pd.DataFrame(breaker_rows)
        df_loads = pd.DataFrame(load_rows)

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df_tree.to_excel(writer, sheet_name="트리구조", index=False)
            df_breakers.to_excel(writer, sheet_name="차단기요약", index=False)
            df_loads.to_excel(writer, sheet_name="하위부하상세", index=False)

            wb = writer.book
            thin = Side(style="thin", color="D1D5DB")

            # -----------------------------
            # 공통 시트 스타일 함수
            # -----------------------------
            def style_sheet(ws):
                header_fill = PatternFill("solid", fgColor="DCE6F8")
                header_font = Font(bold=True, color="1F2937")
                center_align = Alignment(horizontal="center", vertical="center")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)

                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = border

                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.border = border
                        if isinstance(cell.value, (int, float)):
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        else:
                            cell.alignment = Alignment(vertical="center")

                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions

                for column_cells in ws.columns:
                    max_len = 0
                    col_letter = column_cells[0].column_letter
                    for cell in column_cells:
                        val = "" if cell.value is None else str(cell.value)
                        max_len = max(max_len, len(val))
                    ws.column_dimensions[col_letter].width = min(max_len + 3, 32)

            # -----------------------------
            # 트리구조 시트 스타일
            # -----------------------------
            ws_tree = writer.sheets["트리구조"]
            style_sheet(ws_tree)

            breaker_fill = PatternFill("solid", fgColor="EEF4FF")
            load_fill = PatternFill("solid", fgColor="FFF8EF")
            top_fill = PatternFill("solid", fgColor="DCEBFF")

            name_col_idx = None
            level_col_idx = None
            kind_col_idx = None

            headers = [cell.value for cell in ws_tree[1]]
            for i, header in enumerate(headers, start=1):
                if header == "이름":
                    name_col_idx = i
                elif header == "레벨":
                    level_col_idx = i
                elif header == "구분":
                    kind_col_idx = i

            for row_idx in range(2, ws_tree.max_row + 1):
                level_val = ws_tree.cell(row=row_idx, column=level_col_idx).value
                kind_val = ws_tree.cell(row=row_idx, column=kind_col_idx).value

                row_fill = breaker_fill if kind_val == "차단기" else load_fill
                if kind_val == "차단기" and level_val == 0:
                    row_fill = top_fill

                for col_idx in range(1, ws_tree.max_column + 1):
                    cell = ws_tree.cell(row=row_idx, column=col_idx)
                    cell.fill = row_fill

                name_cell = ws_tree.cell(row=row_idx, column=name_col_idx)
                name_cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    indent=int(level_val) * 2 if level_val is not None else 0
                )

                if kind_val == "차단기":
                    name_cell.font = Font(bold=True, color="1E3A8A")
                else:
                    name_cell.font = Font(color="7C2D12")

            ws_tree.column_dimensions["A"].width = 8  # 레벨
            ws_tree.column_dimensions["B"].width = 10  # 구분
            ws_tree.column_dimensions["C"].width = 40  # 이름
            ws_tree.column_dimensions["D"].width = 24  # 상위

            # -----------------------------
            # 차단기요약 / 하위부하상세 시트 스타일
            # -----------------------------
            ws_breakers = writer.sheets["차단기요약"]
            ws_loads = writer.sheets["하위부하상세"]

            style_sheet(ws_breakers)
            style_sheet(ws_loads)

    def _restore_breaker_tree(self, data: dict, parent: Optional[BreakerItem] = None):
        breaker = BreakerItem(
            self,
            data.get("name", self.new_breaker_name(data.get("breaker_type", "MCCB"))),
            float(data.get("x", 120)),
            float(data.get("y", 80)),
            float(data.get("safety_factor", 1.25)),
            parent_breaker=parent,
            is_top_level=bool(data.get("is_top_level", parent is None)),
            width=float(data.get("width", BreakerItem.DEFAULT_WIDTH)),
            height=float(data.get("height", BreakerItem.DEFAULT_HEIGHT)),
            breaker_type=str(data.get("breaker_type", str(data.get("name", "MCCB")).split("-")[0] or "MCCB")),
        )
        self.addItem(breaker)
        self.all_breakers.append(breaker)
        if parent is None:
            self.top_breaker = breaker
        else:
            parent.child_breakers.append(breaker)

        for load in data.get("loads", []):
            part_row = self.db.get_part_by_no(load.get("part_no", ""))
            if part_row is not None:
                breaker.add_part(
                    part_row,
                    int(load.get("quantity", 1)),
                    QPointF(float(load.get("x", breaker.scenePos().x() + 40)),
                            float(load.get("y", breaker.scenePos().y() + 180))),
                    width=float(load.get("width", LoadPartItem.DEFAULT_WIDTH)),
                    height=float(load.get("height", LoadPartItem.DEFAULT_HEIGHT)),
                )

        for child_data in data.get("children", []):
            self._restore_breaker_tree(child_data, breaker)

        breaker.refresh_recursive()
        return breaker

    def load_from_file(self, path: Path):
        self.clear_canvas()
        if not path.exists():
            self.add_top_breaker("MAIN-MCCB", 120, 80, breaker_type="MCCB")
            return
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            top_data = data.get("top_breaker")
            if not top_data:
                self.add_top_breaker("MAIN-MCCB", 120, 80, breaker_type="MCCB")
            else:
                self._restore_breaker_tree(top_data, None)
            self.notify_layout_changed()
        except Exception:
            logger.exception("Failed to load canvas layout")
            self.clear_canvas()
            self.add_top_breaker("MAIN-MCCB", 120, 80, breaker_type="MCCB")


class BreakerCanvasView(QGraphicsView):
    def __init__(self, scene: BreakerCanvasScene):
        super().__init__(scene)
        self._zoom = 1.0
        self._zoom_step = 1.15
        self._zoom_min = 0.35
        self._zoom_max = 3.0
        self.setRenderHint(QPainter.Antialiasing)
        self.setAcceptDrops(True)
        self.setDragMode(QGraphicsView.RubberBandDrag)
        self.setViewportUpdateMode(QGraphicsView.BoundingRectViewportUpdate)
        self.setTransformationAnchor(QGraphicsView.AnchorUnderMouse)
        self.setResizeAnchor(QGraphicsView.AnchorUnderMouse)

    def wheelEvent(self, event: QWheelEvent):
        if event.modifiers() & Qt.ControlModifier:
            factor = self._zoom_step if event.angleDelta().y() > 0 else 1 / self._zoom_step
            next_zoom = self._zoom * factor
            if next_zoom < self._zoom_min:
                factor = self._zoom_min / self._zoom
                self._zoom = self._zoom_min
            elif next_zoom > self._zoom_max:
                factor = self._zoom_max / self._zoom
                self._zoom = self._zoom_max
            else:
                self._zoom = next_zoom
            self.scale(factor, factor)
            event.accept()
            return
        super().wheelEvent(event)

    def reset_zoom(self):
        self.resetTransform()
        self._zoom = 1.0

    def dragEnterEvent(self, event):
        if event.mimeData().hasText():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dragMoveEvent(self, event):
        if event.mimeData().hasText():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event):
        scene_pos = self.mapToScene(event.position().toPoint())
        self.scene().handle_drop(event.mimeData(), scene_pos)
        event.acceptProposedAction()

# ------------------------------------------------------------
# Canvas tab
# ------------------------------------------------------------
class CanvasTab(QWidget):
    def __init__(self, db: DatabaseManager):
        super().__init__()
        self.db = db
        self.scene = BreakerCanvasScene(self.db)
        self._syncing_summary = False
        self.breaker_templates = load_breaker_templates()
        self._build_ui()
        self.scene.load_from_file(CANVAS_SAVE_PATH)
        self.refresh_summary_table()

    def _apply_widget_shadow(self, widget: QWidget):
        effect = QGraphicsDropShadowEffect(widget)
        effect.setBlurRadius(14)
        effect.setOffset(0, 2)
        effect.setColor(QColor(15, 23, 42, 32))
        widget.setGraphicsEffect(effect)

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(10, 10, 10, 10)
        root.setSpacing(10)

        toolbar = QHBoxLayout()
        self.refresh_library_btn = QPushButton("라이브러리 새로고침")
        self.add_top_breaker_btn = QPushButton("최상위 차단기 생성")
        self.main_breaker_type_combo = QComboBox()
        self.main_breaker_type_combo.addItems(["MCCB", "ELCB"])
        self.delete_selected_btn = QPushButton("선택 항목 삭제")
        self.save_canvas_btn = QPushButton("캔버스 저장")
        self.reset_canvas_btn = QPushButton("기본 배치 복원")
        self.reset_zoom_btn = QPushButton("줌 초기화")
        self.toggle_parts_btn = QPushButton("파트")
        self.toggle_parts_btn.setToolTip("파트 라이브러리")
        self.toggle_templates_btn = QPushButton("템플릿")
        self.toggle_templates_btn.setToolTip("차단기 템플릿")
        self.toggle_summary_btn = QPushButton("요약")
        self.toggle_summary_btn.setToolTip("계산 요약")
        for btn in (self.toggle_parts_btn, self.toggle_templates_btn, self.toggle_summary_btn):
            btn.setCheckable(True)
            btn.setChecked(True)
            btn.setProperty("panelToggle", True)
            btn.setMinimumWidth(76)

        toolbar.addWidget(self.refresh_library_btn)
        toolbar.addWidget(QLabel("MAIN 종류"))
        toolbar.addWidget(self.main_breaker_type_combo)
        toolbar.addWidget(self.add_top_breaker_btn)
        toolbar.addWidget(self.delete_selected_btn)
        toolbar.addWidget(self.save_canvas_btn)
        toolbar.addWidget(self.reset_canvas_btn)
        toolbar.addWidget(self.reset_zoom_btn)
        toolbar.addStretch()
        toolbar.addWidget(self.toggle_parts_btn)
        toolbar.addWidget(self.toggle_templates_btn)
        toolbar.addWidget(self.toggle_summary_btn)
        root.addLayout(toolbar)

        self.main_splitter = QSplitter(Qt.Horizontal)
        self.main_splitter.setChildrenCollapsible(True)
        self.main_splitter.setHandleWidth(10)
        root.addWidget(self.main_splitter, 1)

        canvas_group = QGroupBox("단선 구성 캔버스")
        canvas_layout = QVBoxLayout()
        canvas_layout.setContentsMargins(10, 12, 10, 10)
        canvas_layout.setSpacing(8)
        self.canvas_view = BreakerCanvasView(self.scene)
        self.canvas_view.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        canvas_layout.addWidget(self.canvas_view)
        canvas_group.setLayout(canvas_layout)

        self.side_panel = QWidget()
        self.side_panel.setObjectName("canvasSidePanel")
        self.side_panel.setMinimumWidth(280)
        self.side_panel.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        self.side_layout = QVBoxLayout(self.side_panel)
        self.side_layout.setContentsMargins(0, 0, 0, 0)
        self.side_layout.setSpacing(12)

        lib_group = QGroupBox("파트 라이브러리")
        lib_layout = QVBoxLayout()
        lib_layout.setContentsMargins(10, 12, 10, 10)
        lib_layout.setSpacing(8)
        self.part_search_edit = QLineEdit()
        self.part_search_edit.setPlaceholderText("Part No 또는 Part Name 검색")
        self.part_list = DraggablePartListWidget(self.db)
        self.part_list.setMinimumWidth(280)
        lib_layout.addWidget(QLabel("파트를 드래그해서 차단기 위에 놓으면 하위 부하로 추가됩니다."))
        lib_layout.addWidget(self.part_search_edit)
        lib_layout.addWidget(self.part_list)
        lib_group.setLayout(lib_layout)

        breaker_group = QGroupBox("차단기 템플릿")
        breaker_layout = QVBoxLayout()
        breaker_layout.setContentsMargins(10, 12, 10, 10)
        breaker_layout.setSpacing(8)
        self.breaker_template_list = BreakerTemplateListWidget()
        self.breaker_template_list.refresh_templates(self.breaker_templates)
        self.add_template_btn = QPushButton("템플릿 추가")
        breaker_layout.addWidget(QLabel("템플릿을 드래그해서 기존 차단기 위에 놓으세요."))
        breaker_layout.addWidget(self.breaker_template_list)
        breaker_layout.addWidget(self.add_template_btn)
        breaker_group.setLayout(breaker_layout)

        summary_group = QGroupBox("차단기별 계산 요약")
        summary_layout = QVBoxLayout()
        summary_layout.setContentsMargins(10, 12, 10, 10)
        summary_layout.setSpacing(8)
        self.summary_table = QTableWidget()
        self.summary_table.setColumnCount(6)
        self.summary_table.setHorizontalHeaderLabels([
            "차단기", "안전율", "전체 부하 수", "합계 전류(A)", "합계 전력(W)", "추천 차단기(A)"
        ])
        self.summary_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        summary_layout.addWidget(self.summary_table)
        summary_group.setLayout(summary_layout)

        for w in [lib_group, breaker_group, summary_group, canvas_group]:
            self._apply_widget_shadow(w)

        self.main_splitter.addWidget(canvas_group)
        self.main_splitter.addWidget(self.side_panel)
        self.side_layout.addWidget(lib_group)
        self.side_layout.addWidget(breaker_group)
        self.side_layout.addWidget(summary_group)
        self.main_splitter.setSizes([1200, 460])

        self.lib_group = lib_group
        self.breaker_group = breaker_group
        self.summary_group = summary_group
        self._update_side_tab_button(self.toggle_parts_btn, True)
        self._update_side_tab_button(self.toggle_templates_btn, True)
        self._update_side_tab_button(self.toggle_summary_btn, True)
        self._refresh_side_container_layout(initial=True)

        self.refresh_library_btn.clicked.connect(self.reload_library)
        self.add_top_breaker_btn.clicked.connect(self.add_top_breaker)
        self.delete_selected_btn.clicked.connect(self.scene.delete_selected_items)
        self.save_canvas_btn.clicked.connect(self.save_canvas)
        self.reset_canvas_btn.clicked.connect(self.reset_canvas)
        self.reset_zoom_btn.clicked.connect(self.canvas_view.reset_zoom)
        self.toggle_parts_btn.clicked.connect(lambda: self.toggle_side_panel(self.lib_group, self.toggle_parts_btn, "파트"))
        self.toggle_templates_btn.clicked.connect(lambda: self.toggle_side_panel(self.breaker_group, self.toggle_templates_btn, "템플릿"))
        self.toggle_summary_btn.clicked.connect(lambda: self.toggle_side_panel(self.summary_group, self.toggle_summary_btn, "요약"))
        self.add_template_btn.clicked.connect(self.add_breaker_template)
        self.part_search_edit.textChanged.connect(self.on_part_search_changed)
        self.scene.layoutChanged.connect(self.refresh_summary_table)
        self.summary_table.itemChanged.connect(self.on_summary_item_changed)

    def _update_side_tab_button(self, button: QPushButton, active: bool):
        button.blockSignals(True)
        button.setChecked(active)
        button.setProperty("expanded", active)
        button.style().unpolish(button)
        button.style().polish(button)
        button.blockSignals(False)

    def toggle_side_panel(self, widget: QWidget, button: QPushButton, title: str):
        will_show = not widget.isHidden()
        will_show = not will_show
        widget.setHidden(not will_show)
        self._update_side_tab_button(button, will_show)
        self._refresh_side_container_layout()

    def _refresh_side_container_layout(self, initial: bool = False):
        groups = [self.lib_group, self.breaker_group, self.summary_group]
        has_visible = any(not w.isHidden() for w in groups)

        self.side_panel.setVisible(has_visible)
        self.side_panel.updateGeometry()
        self.main_splitter.updateGeometry()

        if has_visible:
            self.side_panel.setMinimumWidth(320)
            self.side_panel.setMaximumWidth(16777215)
            self.side_panel.adjustSize()

            sizes = self.main_splitter.sizes()
            if initial or len(sizes) < 2 or sizes[1] == 0:
                self.main_splitter.setSizes([1200, 460])
            else:
                total = max(sum(sizes), 1400)
                right = max(360, sizes[1])
                left = max(400, total - right)
                self.main_splitter.setSizes([left, right])
        else:
            self.side_panel.setMinimumWidth(0)
            self.side_panel.setMaximumWidth(0)
            total = max(sum(self.main_splitter.sizes()), 1400)
            self.main_splitter.setSizes([total, 0])

    def add_breaker_template(self):
        dialog = BreakerTemplateDialog(self)
        if dialog.exec() != QDialog.Accepted:
            return
        label, prefix = dialog.values()
        if not prefix:
            QMessageBox.warning(self, "오류", "접두어를 입력하세요.")
            return
        if any(t["prefix"] == prefix for t in self.breaker_templates):
            QMessageBox.information(self, "안내", "이미 등록된 템플릿입니다.")
            return
        self.breaker_templates.append({"label": label or prefix, "prefix": prefix})
        save_breaker_templates(self.breaker_templates)
        self.breaker_template_list.refresh_templates(self.breaker_templates)

    def reload_library(self):
        self.part_list.refresh_parts()
        self.part_list.apply_filter(self.part_search_edit.text())
        self.breaker_templates = load_breaker_templates()
        self.breaker_template_list.refresh_templates(self.breaker_templates)

    def on_part_search_changed(self, text: str):
        self.part_list.apply_filter(text)

    def add_top_breaker(self):
        breaker_type = self.main_breaker_type_combo.currentText()
        self.scene.add_top_breaker(f"MAIN-{breaker_type}", 120, 80, breaker_type=breaker_type)

    def save_canvas(self):
        try:
            self.scene.save_to_file(CANVAS_SAVE_PATH)
            self.scene.export_to_excel(CANVAS_EXPORT_PATH)
            QMessageBox.information(
                self,
                "완료",
                f"캔버스 구성이 저장되었습니다.\nJSON: {CANVAS_SAVE_PATH}\nExcel: {CANVAS_EXPORT_PATH}"
            )
        except Exception as e:
            logger.exception("Failed to save canvas")
            QMessageBox.critical(self, "오류", str(e))

    def reset_canvas(self):
        breaker_type = self.main_breaker_type_combo.currentText()
        self.scene.clear_canvas()
        self.scene.add_top_breaker(f"MAIN-{breaker_type}", 120, 80, breaker_type=breaker_type)
        self.scene.notify_layout_changed()
        self.canvas_view.reset_zoom()
        self.save_canvas()

    def _collect_breakers(self, root_breaker: Optional[BreakerItem]) -> List[BreakerItem]:
        if root_breaker is None:
            return []
        result = []

        def walk(b: BreakerItem):
            result.append(b)
            for child in b.child_breakers:
                walk(child)

        walk(root_breaker)
        return result

    def refresh_summary_table(self):
        if self.scene.top_breaker is not None:
            top_type = (self.scene.top_breaker.breaker_type or "MCCB").upper()
            idx = self.main_breaker_type_combo.findText(top_type)
            if idx >= 0:
                self.main_breaker_type_combo.setCurrentIndex(idx)

        breakers = self._collect_breakers(self.scene.top_breaker)
        self._syncing_summary = True
        try:
            self.summary_table.setRowCount(len(breakers))
            for row, breaker in enumerate(breakers):
                values = [
                    breaker.name,
                    f"{breaker.safety_factor:.2f}",
                    str(breaker.get_total_load_count()),
                    f"{breaker.get_total_current():.2f}",
                    f"{breaker.get_total_power():.0f}",
                    str(breaker.suggested_breaker()),
                ]
                for col, value in enumerate(values):
                    item = self.summary_table.item(row, col)
                    if item is None:
                        item = QTableWidgetItem()
                        self.summary_table.setItem(row, col, item)
                    item.setText(value)
                    item.setData(Qt.UserRole, breaker)
                    flags = Qt.ItemIsSelectable | Qt.ItemIsEnabled
                    if col == 0:
                        flags |= Qt.ItemIsEditable
                    item.setFlags(flags)
        finally:
            self._syncing_summary = False

    def on_summary_item_changed(self, item: QTableWidgetItem):
        if self._syncing_summary or item.column() != 0:
            return
        breaker = item.data(Qt.UserRole)
        if not isinstance(breaker, BreakerItem):
            return
        new_name = item.text().strip()
        if not new_name:
            self._syncing_summary = True
            try:
                item.setText(breaker.name)
            finally:
                self._syncing_summary = False
            return
        if breaker.name != new_name:
            breaker.name = new_name
            breaker.refresh_recursive()
            self.scene.notify_layout_changed()

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Electrical Capacity & Breaker Helper")
        self.resize(1680, 980)
        self.db = DatabaseManager(DB_PATH)
        self._build_ui()

    def _build_ui(self):
        tabs = QTabWidget()
        self.parts_tab = PartsTab(self.db)
        self.calc_tab = CalcTab(self.db)
        self.canvas_tab = CanvasTab(self.db)
        tabs.addTab(self.parts_tab, "파트 관리")
        tabs.addTab(self.calc_tab, "부하 계산")
        tabs.addTab(self.canvas_tab, "차단기 캔버스")
        self.setCentralWidget(tabs)

        refresh_action = QAction("새로고침", self)
        refresh_action.triggered.connect(self.refresh_all)
        self.menuBar().addAction(refresh_action)

    def refresh_all(self):
        self.parts_tab.refresh_table()
        self.calc_tab.reload_part_numbers()
        self.canvas_tab.reload_library()
        self.canvas_tab.refresh_summary_table()

    def closeEvent(self, event):
        try:
            self.canvas_tab.scene.save_to_file(CANVAS_SAVE_PATH)
        except Exception:
            logger.exception("Failed to save canvas on close")
        super().closeEvent(event)


def load_stylesheet(app: QApplication):
    if STYLE_PATH.exists():
        app.setStyleSheet(STYLE_PATH.read_text(encoding="utf-8"))
    else:
        logger.warning("Stylesheet file not found: %s", STYLE_PATH)


def main():
    app = QApplication(sys.argv)
    load_stylesheet(app)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
